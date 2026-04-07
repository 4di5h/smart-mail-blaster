from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import smtplib
import csv
import time
import subprocess
import os
import io
import json
import tempfile
import re
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook

EMAIL_REGEX = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"

app = Flask(__name__, static_folder=".")
CORS(app)

CRED_PATH = os.path.expandvars(r"%USERPROFILE%\hashed_smtp_seb.xml")
ATTACHMENTS_DIR = os.path.join(os.path.dirname(__file__), "attachments")

os.makedirs(ATTACHMENTS_DIR, exist_ok=True)


def encrypt_to_xml(plaintext, path):
    with tempfile.NamedTemporaryFile(delete=False, mode="w", encoding="utf-8") as tmp:
        tmp.write(plaintext)
        tmp_path = tmp.name

    ps_cmd = f"""
    $text = Get-Content "{tmp_path}" -Raw
    $secure = ConvertTo-SecureString $text -AsPlainText -Force
    $encrypted = $secure | ConvertFrom-SecureString
    Set-Content -Path "{path}" -Value $encrypted -NoNewline
    """

    subprocess.run(["powershell", "-Command", ps_cmd], check=True)

    os.remove(tmp_path)


def load_credentials():
    ps_cmd = (
        f'$p = Get-Content "{CRED_PATH}" -Raw | ConvertTo-SecureString; '
        f'[Runtime.InteropServices.Marshal]::PtrToStringAuto('
        f'[Runtime.InteropServices.Marshal]::SecureStringToBSTR($p))'
    )

    result = subprocess.run(
        ["powershell", "-Command", ps_cmd],
        capture_output=True, text=True
    )

    output = result.stdout.strip()

    if not output:
        return None

    return json.loads(output)

def load_credentials_safe():
    try:
        creds = load_credentials()
        if not creds or "user" not in creds or "pass" not in creds:
            return None
        return creds
    except Exception:
        return None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/api/status", methods=["GET"])
def status():

    # Step 1: Check if file exists
    if not os.path.exists(CRED_PATH):
        return jsonify({"configured": False})

    # Step 2: Try loading credentials
    try:
        creds = load_credentials()
        if creds and "user" in creds:
            return jsonify({
                "configured": True,
                "user": creds["user"]
            })
    except Exception:
        pass

    # Step 3: File exists but broken
    return jsonify({
        "configured": False,
        "error": "Credential file exists but could not be read"
    })

@app.route("/api/setup", methods=["POST"])
def setup():
    data = request.json
    email = data.get("email", "").strip()
    password = data.get("password", "").strip()
    if not email or not password:
        return jsonify({"success": False, "error": "Email and password required"}), 400
    try:
        payload = json.dumps({"user": email, "pass": password})
        encrypt_to_xml(payload, CRED_PATH)
        return jsonify({"success": True, "user": email})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/attachments", methods=["GET"])
def list_attachments():
    files = os.listdir(ATTACHMENTS_DIR)
    return jsonify({"files": files})


@app.route("/api/send", methods=["POST"])
def send_emails():

    template = request.form.get("template", "").strip()
    subject = request.form.get("subject", "").strip()
    csv_file = request.files.get("csv")
    delay = int(request.form.get("delay", 10))

    if not template or not subject or not csv_file:
        return jsonify({
            "success": False,
            "error": "Template, subject, and CSV are required"
        }), 400

    creds = load_credentials_safe()
    if not creds:
        return jsonify({
            "success": False,
            "error": "Credentials not set. Please run setup first."
        }), 400

    gmail_user = creds["user"]
    gmail_pass = creds["pass"]

    file = request.files.get("csv")

    filename = file.filename.lower()

    if filename.endswith(".csv"):
        content = file.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)[1:]

    elif filename.endswith(".xlsx"):
        wb = load_workbook(file, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))

    else:
        return jsonify({"success": False, "error": "Unsupported file format"}), 400

    attachment_files = os.listdir(ATTACHMENTS_DIR)

    results = []

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(gmail_user, gmail_pass)

            for i, row in enumerate(rows):
                results.append({
                    "index": i,
                    "status": "skipped"
                })
                    continue

                name = str(row[0]).strip()
                email = str(row[1]).strip()

                if not name or not email:
                    continue

                # Add this validation
                if not re.match(EMAIL_REGEX, email):
                    return jsonify({
                        "success": False,
                        "error": f"Invalid email detected: {email}"
                    }), 400

                try:
                    msg = MIMEMultipart()
                    msg["From"] = gmail_user
                    msg["To"] = email
                    msg["Subject"] = subject
                    msg.attach(MIMEText(template.replace("{referenceName}", name), "plain"))

                    for fname in attachment_files:
                        fpath = os.path.join(ATTACHMENTS_DIR, fname)
                        with open(fpath, "rb") as f:
                            part = MIMEBase("application", "octet-stream")
                            part.set_payload(f.read())
                            encoders.encode_base64(part)
                            part.add_header("Content-Disposition", f"attachment; filename={fname}")
                            msg.attach(part)

                    smtp.sendmail(gmail_user, email, msg.as_string())
                    results.append({"index": i, "email": email, "name": name, "status": "sent"})
                    time.sleep(delay)

                except Exception as e:
                    results.append({"index": i, "email": email, "name": name, "status": f"failed: {str(e)}"})

    except Exception as e:
        return jsonify({"success": False, "error": f"Auth failed: {str(e)}"}), 500

    return jsonify({"success": True, "results": results})


if __name__ == "__main__":
    app.run(debug=True, port=6001)
